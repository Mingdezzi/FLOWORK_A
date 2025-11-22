import pandas as pd
import numpy as np
from openpyxl.utils import column_index_from_string
from flowork.models import db, Product, Variant, StoreStock, Setting, Store
from flowork.utils import clean_string_upper, get_choseong, generate_barcode
import traceback
import json
from sqlalchemy.orm import selectinload
from sqlalchemy.exc import IntegrityError

try:
    from flowork.services.transformer import transform_horizontal_to_vertical
except ImportError:
    transform_horizontal_to_vertical = None

def _get_column_indices_from_form(form, field_map, strict=True):
    column_map_indices = {}
    missing_fields = []
    
    for field_name, (form_key, is_required) in field_map.items():
        col_letter = form.get(form_key)
        
        if strict and is_required and not col_letter:
            missing_fields.append(field_name)
        
        if col_letter:
            try:
                column_map_indices[field_name] = column_index_from_string(col_letter) - 1
            except ValueError:
                column_map_indices[field_name] = None
        else:
            column_map_indices[field_name] = None

    if missing_fields:
        raise ValueError(f"다음 필수 항목의 열이 선택되지 않았습니다: {', '.join(missing_fields)}")
            
    return column_map_indices

def _read_excel_data_to_df(file_stream, column_map_indices):
    try:
        if hasattr(file_stream, 'seek'):
            file_stream.seek(0)
        df = pd.read_excel(file_stream, header=0)
    except Exception:
        if hasattr(file_stream, 'seek'):
            file_stream.seek(0)
        try:
            df = pd.read_csv(file_stream)
        except:
            return pd.DataFrame()

    if df.empty:
        return pd.DataFrame()

    selected_cols = {}
    total_cols = df.shape[1]
    for field_name, col_idx in column_map_indices.items():
        if col_idx is not None and 0 <= col_idx < total_cols:
            original_col_name = df.columns[col_idx]
            selected_cols[original_col_name] = field_name
    
    if not selected_cols:
        return pd.DataFrame()

    df_subset = df[list(selected_cols.keys())].rename(columns=selected_cols)
    
    for field in column_map_indices.keys():
        if field not in df_subset.columns:
            df_subset[field] = np.nan
            
    return df_subset

def _optimize_dataframe(df, brand_settings, upload_mode):
    if df.empty: return df

    required = ['product_number', 'color', 'size']
    df = df.dropna(subset=required)
    
    str_cols = ['product_number', 'product_name', 'color', 'size', 'item_category']
    for col in str_cols:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip().replace({'nan': None, 'None': None})

    num_cols = ['original_price', 'sale_price', 'release_year', 'hq_stock', 'store_stock']
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)

    if 'original_price' in df.columns and 'sale_price' in df.columns:
        op = df['original_price']
        sp = df['sale_price']
        df['sale_price'] = np.where((op > 0) & (sp == 0), op, sp)
        df['original_price'] = np.where((sp > 0) & (op == 0), sp, op)

    if 'barcode' not in df.columns: df['barcode'] = None
    
    mask_no_barcode = df['barcode'].isna() | (df['barcode'] == '')
    if mask_no_barcode.any():
        df.loc[mask_no_barcode, 'barcode'] = df[mask_no_barcode].apply(
            lambda row: generate_barcode(row.to_dict(), brand_settings), axis=1
        )

    df = df.dropna(subset=['barcode'])
    
    df['product_number_cleaned'] = df['product_number'].apply(clean_string_upper)
    df['barcode_cleaned'] = df['barcode'].apply(clean_string_upper)
    
    if 'color' in df.columns:
        df['color_cleaned'] = df['color'].apply(clean_string_upper)
    if 'size' in df.columns:
        df['size_cleaned'] = df['size'].apply(clean_string_upper)
    
    if 'product_name' in df.columns:
        df['product_name_cleaned'] = df['product_name'].apply(clean_string_upper)
        df['product_name_choseong'] = df['product_name'].apply(get_choseong)
    
    if 'is_favorite' not in df.columns:
        df['is_favorite'] = 0
    else:
        df['is_favorite'] = pd.to_numeric(df['is_favorite'], errors='coerce').fillna(0).astype(int)
    
    df = df.drop_duplicates(subset=['barcode_cleaned'], keep='last')

    return df

def verify_stock_excel(file_path, form, upload_mode):
    field_map = {'product_number': ('col_pn', True)}
    
    try:
        column_map_indices = _get_column_indices_from_form(form, field_map, strict=False)
        
        with open(file_path, 'rb') as f:
            df = _read_excel_data_to_df(f, column_map_indices)
        
        if df.empty:
            return {'status': 'success', 'suspicious_rows': []}

        df['_row_index'] = df.index + 2
        suspicious_rows = []
        
        for _, row in df.iterrows():
            pn = row.get('product_number')
            if pd.isna(pn) or str(pn).strip() == "":
                suspicious_rows.append({
                    'row_index': int(row['_row_index']), 
                    'preview': '(품번없음)', 
                    'reasons': '품번 누락'
                })
                
        return {'status': 'success', 'suspicious_rows': suspicious_rows[:100]}

    except Exception as e:
        return {'status': 'error', 'message': f"검증 중 오류: {e}"}

def import_excel_file(file, form, brand_id, progress_callback=None):
    if not file: return False, '파일이 없습니다.', 'error'
    BATCH_SIZE = 5000  # [최적화] 12GB RAM 활용을 위해 배치 사이즈 증가
    
    is_horizontal = form.get('is_horizontal') == 'on'
    
    try:
        settings_query = Setting.query.filter_by(brand_id=brand_id).all()
        brand_settings = {s.key: s.value for s in settings_query}
        
        field_map = {
            'product_number': ('col_pn', True),
            'color': ('col_color', True),
            'product_name': ('col_pname', False),
            'release_year': ('col_year', False),
            'item_category': ('col_category', False),
            'original_price': ('col_oprice', False),
            'sale_price': ('col_sprice', False),
            'is_favorite': ('col_favorite', False)
        }

        if is_horizontal:
            import_strategy = 'horizontal_matrix'
        else:
            import_strategy = None
            field_map['size'] = ('col_size', True)
            
        column_map_indices = _get_column_indices_from_form(form, field_map, strict=False)

        df = pd.DataFrame()
        if import_strategy == 'horizontal_matrix':
            if transform_horizontal_to_vertical is None:
                return False, 'pandas 라이브러리가 필요합니다.', 'error'
            try:
                size_mapping_config = json.loads(brand_settings.get('SIZE_MAPPING', '{}'))
                category_mapping_config = json.loads(brand_settings.get('CATEGORY_MAPPING_RULE', '{}'))
            except json.JSONDecodeError:
                return False, '브랜드 설정 형식이 잘못되었습니다.', 'error'
            
            df = transform_horizontal_to_vertical(file, size_mapping_config, category_mapping_config, column_map_indices)
        else:
            df = _read_excel_data_to_df(file, column_map_indices)

        df = _optimize_dataframe(df, brand_settings, 'db')
        
        if df.empty:
            return False, "유효한 데이터가 없습니다.", 'error'
            
        # [수정] 기존 데이터 삭제 시 무결성 오류 처리 추가
        try:
            store_ids = db.session.query(Store.id).filter_by(brand_id=brand_id)
            db.session.query(StoreStock).filter(StoreStock.store_id.in_(store_ids)).delete(synchronize_session=False)
            product_ids = db.session.query(Product.id).filter_by(brand_id=brand_id)
            db.session.query(Variant).filter(Variant.product_id.in_(product_ids)).delete(synchronize_session=False)
            db.session.query(Product).filter_by(brand_id=brand_id).delete(synchronize_session=False)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return False, "삭제 실패: 판매 내역이나 주문 내역에 사용 중인 상품이 있습니다. 시스템 관리에서 판매/주문 데이터를 먼저 초기화해주세요.", 'error'

        products_id_map = {} 
        total_products = 0
        total_variants = 0
        
        records = df.to_dict('records')
        total_items = len(records)

        for i in range(0, total_items, BATCH_SIZE):
            if progress_callback: progress_callback(i, total_items)
            batch = records[i:i+BATCH_SIZE]
            
            products_to_add = []
            new_products_in_batch = {} 

            for item in batch:
                pn = item['product_number_cleaned']
                if pn not in products_id_map and pn not in new_products_in_batch:
                    
                    pname = item.get('product_name') or item['product_number']
                    pn_cleaned_val = item.get('product_name_cleaned') or clean_string_upper(pname)
                    choseong_val = item.get('product_name_choseong') or get_choseong(pname)
                    
                    p = Product(
                        brand_id=brand_id,
                        product_number=item['product_number'],
                        product_name=pname,
                        release_year=item['release_year'] if item['release_year'] > 0 else None,
                        item_category=item['item_category'],
                        is_favorite=item.get('is_favorite', 0),
                        product_number_cleaned=pn,
                        product_name_cleaned=pn_cleaned_val,
                        product_name_choseong=choseong_val
                    )
                    new_products_in_batch[pn] = p
                    products_to_add.append(p)
            
            if products_to_add:
                db.session.add_all(products_to_add)
                db.session.flush()
                for pn, p_obj in new_products_in_batch.items():
                    products_id_map[pn] = p_obj.id
                total_products += len(products_to_add)
            
            variants_to_add = []
            for item in batch:
                pid = products_id_map.get(item['product_number_cleaned'])
                if pid:
                    v = Variant(
                        product_id=pid,
                        barcode=item['barcode'],
                        color=item['color'],
                        size=item['size'],
                        original_price=item['original_price'],
                        sale_price=item['sale_price'],
                        hq_quantity=item.get('hq_stock', 0), 
                        barcode_cleaned=item['barcode_cleaned'],
                        color_cleaned=item.get('color_cleaned', clean_string_upper(item['color'])),
                        size_cleaned=item.get('size_cleaned', clean_string_upper(item['size']))
                    )
                    variants_to_add.append(v)
            
            if variants_to_add:
                db.session.bulk_save_objects(variants_to_add)
                total_variants += len(variants_to_add)
            
            db.session.commit()
            
        if progress_callback: progress_callback(total_items, total_items)
        return True, f"초기화 완료: 상품 {total_products}개, 옵션 {total_variants}개", 'success'
        
    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return False, f"오류: {e}", 'error'

def process_stock_upsert_excel(file_path, form, upload_mode, brand_id, target_store_id=None, progress_callback=None, excluded_row_indices=None, allow_create=True):
    try:
        settings_query = Setting.query.filter_by(brand_id=brand_id).all()
        brand_settings = {s.key: s.value for s in settings_query}
        
        is_horizontal = form.get('is_horizontal') == 'on'

        field_map = {
            'product_number': ('col_pn', True),
            'color': ('col_color', True),
            'product_name': ('col_pname', False),
            'release_year': ('col_year', False),
            'item_category': ('col_category', False),
            'original_price': ('col_oprice', False),
            'sale_price': ('col_sprice', False),
            'is_favorite': ('col_favorite', False)
        }
        
        import_strategy = None

        if upload_mode == 'hq':
            if is_horizontal:
                import_strategy = 'horizontal_matrix'
            else:
                field_map['size'] = ('col_size', True)
                field_map['hq_stock'] = ('col_hq_stock', True)

        elif upload_mode == 'store':
            if not target_store_id: return 0, 0, '매장 ID 누락', 'error'
            if is_horizontal:
                import_strategy = 'horizontal_matrix'
            else:
                field_map['size'] = ('col_size', True)
                field_map['store_stock'] = ('col_store_stock', True)
            
        column_map_indices = _get_column_indices_from_form(form, field_map, strict=False)

        with open(file_path, 'rb') as f:
            df = pd.DataFrame()
            if import_strategy == 'horizontal_matrix' and transform_horizontal_to_vertical:
                try:
                    size_conf = json.loads(brand_settings.get('SIZE_MAPPING', '{}'))
                    cat_conf = json.loads(brand_settings.get('CATEGORY_MAPPING_RULE', '{}'))
                    df = transform_horizontal_to_vertical(f, size_conf, cat_conf, column_map_indices)
                    
                    if upload_mode == 'store' and 'hq_stock' in df.columns:
                        df.rename(columns={'hq_stock': 'store_stock'}, inplace=True)
                        
                except Exception as e:
                    return 0, 0, f"매트릭스 변환 오류: {e}", 'error'
            else:
                df = _read_excel_data_to_df(f, column_map_indices)
            
        if df.empty:
            return 0, 0, "처리할 데이터가 없습니다.", "warning"
            
        if excluded_row_indices:
            if '_row_index' in df.columns:
                df = df[~df['_row_index'].isin(excluded_row_indices)]

        df = _optimize_dataframe(df, brand_settings, upload_mode)
        
        if df.empty: return 0, 0, "유효한 데이터 없음", "warning"

        pn_list = df['product_number_cleaned'].unique().tolist()
        
        products_in_db = Product.query.filter(
            Product.brand_id == brand_id, 
            Product.product_number_cleaned.in_(pn_list)
        ).options(
            selectinload(Product.variants).selectinload(Variant.stock_levels)
        ).all()
        
        product_map = {p.product_number_cleaned: p for p in products_in_db}
        
        variant_map = {} 
        for p in products_in_db:
            for v in p.variants: variant_map[v.barcode_cleaned] = v
        
        store_stock_map = {}
        if upload_mode == 'store':
            v_ids = [v.id for v in variant_map.values()]
            if v_ids:
                stocks = db.session.query(StoreStock).filter(StoreStock.store_id==target_store_id, StoreStock.variant_id.in_(v_ids)).all()
                store_stock_map = {s.variant_id: s for s in stocks}

        cnt_prod = 0; cnt_var = 0; cnt_update = 0
        
        records = df.to_dict('records')
        total_items = len(records)

        for idx, item in enumerate(records):
            if progress_callback and idx % 500 == 0: progress_callback(idx, total_items)
            
            try:
                pn_clean = item['product_number_cleaned']
                bc_clean = item['barcode_cleaned']

                prod = product_map.get(pn_clean)
                if not prod:
                    if not allow_create: continue
                    pname = item.get('product_name') or item['product_number']
                    
                    pn_cleaned_val = item.get('product_name_cleaned') or clean_string_upper(pname)
                    choseong_val = item.get('product_name_choseong') or get_choseong(pname)

                    prod = Product(
                        brand_id=brand_id, 
                        product_number=item['product_number'], 
                        product_name=pname, 
                        product_number_cleaned=pn_clean, 
                        product_name_cleaned=pn_cleaned_val, 
                        product_name_choseong=choseong_val
                    )
                    db.session.add(prod)
                    product_map[pn_clean] = prod
                    cnt_prod += 1
                
                if item.get('release_year') and item['release_year'] > 0: 
                    prod.release_year = item['release_year']
                if item.get('item_category'): 
                    prod.item_category = item['item_category']
                if item.get('is_favorite') == 1:
                    prod.is_favorite = 1
                
                var = variant_map.get(bc_clean)
                op = item['original_price']
                sp = item['sale_price']

                if not var:
                    if not allow_create: continue
                    var = Variant(
                        product=prod, 
                        barcode=item['barcode'], 
                        color=item['color'], 
                        size=item['size'], 
                        original_price=op, 
                        sale_price=sp, 
                        hq_quantity=0, 
                        barcode_cleaned=bc_clean,
                        color_cleaned=item.get('color_cleaned', clean_string_upper(item['color'])),
                        size_cleaned=item.get('size_cleaned', clean_string_upper(item['size']))
                    )
                    db.session.add(var)
                    variant_map[bc_clean] = var
                    cnt_var += 1
                else:
                    if op > 0: var.original_price = op
                    if sp > 0: var.sale_price = sp
                
                if upload_mode == 'hq' and 'hq_stock' in item:
                    var.hq_quantity = item['hq_stock']
                    cnt_update += 1
                
                elif upload_mode == 'store' and 'store_stock' in item:
                    qty = item['store_stock']
                    
                    if var.id and var.id in store_stock_map:
                        store_stock_map[var.id].quantity = qty
                        cnt_update += 1
                    else:
                        found = False
                        if hasattr(var, 'stock_levels'):
                            for s in var.stock_levels:
                                if s.store_id == target_store_id:
                                    s.quantity = qty
                                    found = True
                                    break
                        
                        if not found:
                            new_stk = StoreStock(store_id=target_store_id, quantity=qty)
                            var.stock_levels.append(new_stk)
                        
                        cnt_update += 1

            except: continue

        db.session.commit()
        
        if progress_callback: progress_callback(total_items, total_items)
        return cnt_update, cnt_var, f"완료: 상품 {cnt_prod} / 옵션 {cnt_var} 생성, {cnt_update}건 업데이트", 'success'

    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return 0, 0, f"오류: {e}", 'error'

def _process_stock_update_excel(file, form, upload_mode, brand_id, target_store_id):
    try:
        field_map = {
            'barcode': ('barcode_col', True),
            'qty': ('qty_col', True)
        }
        column_map_indices = _get_column_indices_from_form(form, field_map)
        
        with open(file, 'rb') as f:
            df = _read_excel_data_to_df(f, column_map_indices)
            
        if df.empty: return 0, 0, "데이터 없음", "warning"

        df['barcode_cleaned'] = df['barcode'].astype(str).apply(clean_string_upper)
        df['qty'] = pd.to_numeric(df['qty'], errors='coerce').fillna(0).astype(int)
        
        barcode_qty_map = dict(zip(df['barcode_cleaned'], df['qty']))
        
        if not barcode_qty_map:
            return 0, 0, "유효한 데이터가 없습니다.", "warning"
            
        variants = db.session.query(Variant).join(Product).filter(
            Product.brand_id == brand_id,
            Variant.barcode_cleaned.in_(barcode_qty_map.keys())
        ).all()
        
        updated_count = 0
        
        if upload_mode == 'hq':
            for v in variants:
                v.hq_quantity = barcode_qty_map[v.barcode_cleaned]
                updated_count += 1
        elif upload_mode == 'store':
            variant_ids = [v.id for v in variants]
            existing_stocks = db.session.query(StoreStock).filter(
                StoreStock.store_id == target_store_id,
                StoreStock.variant_id.in_(variant_ids)
            ).all()
            stock_map = {s.variant_id: s for s in existing_stocks}
            
            for v in variants:
                new_qty = barcode_qty_map[v.barcode_cleaned]
                if v.id in stock_map:
                    stock_map[v.id].quantity = new_qty
                else:
                    new_stock = StoreStock(store_id=target_store_id, variant_id=v.id, quantity=new_qty)
                    db.session.add(new_stock)
                updated_count += 1
                
        db.session.commit()
        return updated_count, 0, f"{updated_count}건의 재고가 업데이트되었습니다.", "success"

    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return 0, 0, f"오류 발생: {e}", "error"

def export_db_to_excel(brand_id):
    import io
    import openpyxl
    from datetime import datetime
    try:
        query = db.session.query(
            Product.product_number, Product.product_name, Product.release_year, Product.item_category,
            Variant.barcode, Variant.color, Variant.size, Variant.original_price, Variant.sale_price, Variant.hq_quantity,
            Product.is_favorite
        ).join(Variant, Product.id == Variant.product_id).filter(Product.brand_id == brand_id)
        
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet()
        ws.append(["품번", "품명", "연도", "카테고리", "바코드", "컬러", "사이즈", "정상가", "판매가", "본사재고", "즐겨찾기"])
        
        for row in query.yield_per(1000):
            ws.append(list(row))
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, f"db_backup_{datetime.now().strftime('%Y%m%d')}.xlsx", None
    except Exception as e:
        print(f"Export Error: {e}")
        traceback.print_exc()
        return None, None, str(e)

def export_stock_check_excel(store_id, brand_id):
    import io
    import openpyxl
    from datetime import datetime
    try:
        variants = db.session.query(Variant).join(Product).filter(Product.brand_id == brand_id).all()
        stocks = db.session.query(StoreStock).filter_by(store_id=store_id).all()
        stock_map = {s.variant_id: s for s in stocks}
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["품번", "품명", "컬러", "사이즈", "바코드", "전산재고", "실사재고", "차이"])
        
        for v in variants:
            st = stock_map.get(v.id)
            qty = st.quantity if st else 0
            actual = st.actual_stock if st and st.actual_stock is not None else ''
            diff = (qty - actual) if isinstance(actual, int) else ''
            ws.append([v.product.product_number, v.product.product_name, v.color, v.size, v.barcode, qty, actual, diff])
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, f"stock_check_{datetime.now().strftime('%Y%m%d')}.xlsx", None
    except Exception as e:
        return None, None, str(e)