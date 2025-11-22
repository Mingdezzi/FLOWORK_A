import io
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime
from flowork.models import db, Product, Variant, StoreStock

def export_db_to_excel(brand_id):
    try:
        query = db.session.query(
            Product.product_number, Product.product_name, Product.release_year, Product.item_category,
            Variant.barcode, Variant.color, Variant.size, Variant.original_price, Variant.sale_price, Variant.hq_quantity,
            Product.is_favorite
        ).join(Variant, Product.id == Variant.product_id).filter(Product.brand_id == brand_id)
        
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet()
        ws.append(["품번", "품명", "연도", "카테고리", "바코드", "컬러", "사이즈", "정상가", "판매가", "본사재고", "즐겨찾기"])
        
        for row in query.yield_per(1000):
            ws.append(list(row))
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, f"db_backup_{datetime.now().strftime('%Y%m%d')}.xlsx", None
    except Exception as e:
        return None, None, str(e)

def export_stock_check_excel(store_id, brand_id):
    try:
        variants = db.session.query(Variant).join(Product).filter(Product.brand_id == brand_id).all()
        stocks = db.session.query(StoreStock).filter_by(store_id=store_id).all()
        stock_map = {s.variant_id: s for s in stocks}
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["품번", "품명", "컬러", "사이즈", "바코드", "전산재고", "실사재고", "차이"])
        
        for v in variants:
            st = stock_map.get(v.id)
            qty = st.quantity if st else 0
            actual = st.actual_stock if st and st.actual_stock is not None else ''
            diff = (qty - actual) if isinstance(actual, int) else ''
            ws.append([v.product.product_number, v.product.product_name, v.color, v.size, v.barcode, qty, actual, diff])
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, f"stock_check_{datetime.now().strftime('%Y%m%d')}.xlsx", None
    except Exception as e:
        return None, None, str(e)

def analyze_excel_file(file_storage):
    try:
        file_bytes = file_storage.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        
        max_col = min(ws.max_column, 26)
        letters = [get_column_letter(i) for i in range(1, max_col + 1)]
        
        preview = {}
        max_row = min(6, ws.max_row + 1)
        
        if max_row <= 1: return None, "데이터가 없습니다."

        for letter in letters:
            col_idx = column_index_from_string(letter)
            col_data = []
            for i in range(1, max_row):
                val = ws.cell(row=i, column=col_idx).value
                col_data.append(str(val) if val is not None else '')
            preview[letter] = col_data
            
        return {'letters': letters, 'preview': preview}, None
    except Exception as e:
        return None, str(e)